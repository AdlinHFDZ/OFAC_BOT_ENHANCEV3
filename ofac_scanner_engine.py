"""
ofac_scanner_engine.py
Orchestrates one scan run: reads Excel/text/archive files from disk, runs
detection + extraction (ofac_extraction_engine) on each, writes per-file/
per-sheet results to the database (ofac_database), and archives each file
after it's successfully processed.

TEST COVERAGE NOTE: verified end to end on real Windows with real polars/
openpyxl/msoffcrypto/7-Zip via test_scanner_engine.py -- run that file (or
the whole test suite) after any change here, especially to the chunked-
reading paths, since that's where both critical big-file bugs from the
original code review lived, and the file most worth re-verifying first if
anything regresses.

No DPAPI/Windows-specific dependency remains in this file -- the password
vault dropped DPAPI encryption (see ofac_password_vault.py's docstring for
why), and msoffcrypto's Excel decryption itself works cross-platform.

BUG FIXES from the review, specifically:
  1. Text file chunking (`for batch in reader:` on a BatchedCsvReader, which
     doesn't support iteration) is replaced with the documented
     next_batches(n)-in-a-loop pattern, re-called every iteration.
  2. Excel files now actually chunk above EXCEL_SIZE_TO_CHUNK, via a
     streaming read (openpyxl read_only mode) instead of always loading the
     whole workbook into memory. This didn't exist at all before (v2 had a
     hardcoded `use_chunking = False` and a stub that raised
     NotImplementedError).
Both paths use the same "detect columns once on the first chunk, reuse that
mapping for every later chunk" approach, since re-running detection per
chunk would be wrong (later chunks have no header row) and wasteful.
"""

import os
import io
import time
import subprocess
from dataclasses import dataclass, field
from datetime import datetime

import polars as pl
import msoffcrypto
import openpyxl

from ofac_constants import (
    FILE_EXTENSIONS_EXCEL, FILE_EXTENSIONS_TEXT, FILE_EXTENSIONS_ARCHIVE,
    TEXT_SIZE_TO_CHUNK, EXCEL_SIZE_TO_CHUNK, CHUNK_ROWS, MAX_SEARCH_ROWS,
    MAX_ROWS_PER_OUTPUT_CSV, SEVEN_ZIP_PATH,
)
from ofac_file_utils import get_unique_save_path, move_file_to_archived, get_all_files, detect_encoding_and_dialect
from ofac_password_retry import try_passwords
from ofac_password_vault import mask_password
import ofac_header_detection as detection
import ofac_extraction_engine as engine
import ofac_database as db


# ==================== SCAN JOB ====================

@dataclass
class ScanJob:
    run_id: int
    company_code: str
    csv_folder: str
    archived_folder: str
    unzipped_folder: str
    passwords: list  # list of (label, password) tuples, already resolved from the vault
    stop_flag: callable = field(default=lambda: False)  # returns True if the user hit Stop


# ==================== OUTPUT PART-FILE WRITER ====================

class PartedCSVWriter:
    """
    Accumulates polars DataFrames and flushes to numbered output CSV files
    (`..._part1_OFAC_OUTPUT.csv`, `..._part2_...`, ...) once
    MAX_ROWS_PER_OUTPUT_CSV is reached, so no single output file exceeds the
    configured row limit -- replaces v2's fragile comma-joined-path approach
    for tracking multi-part output with a clean list of (path, row_count).
    """

    def __init__(self, base_path_no_ext, max_rows_per_file=MAX_ROWS_PER_OUTPUT_CSV):
        self.base_path_no_ext = base_path_no_ext
        self.max_rows_per_file = max_rows_per_file
        self.part_number = 1
        self._buffer = []
        self._buffer_rows = 0
        self.written_parts = []  # list of (path, row_count)

    def add(self, df_chunk):
        if df_chunk.height == 0:
            return
        self._buffer.append(df_chunk)
        self._buffer_rows += df_chunk.height
        while self._buffer_rows >= self.max_rows_per_file:
            self._flush(up_to_rows=self.max_rows_per_file)

    def _flush(self, up_to_rows=None):
        if not self._buffer:
            return
        combined = pl.concat(self._buffer, how="vertical") if len(self._buffer) > 1 else self._buffer[0]

        if up_to_rows is not None and combined.height > up_to_rows:
            to_write = combined.slice(0, up_to_rows)
            remainder = combined.slice(up_to_rows)
            self._buffer = [remainder]
            self._buffer_rows = remainder.height
        else:
            to_write = combined
            self._buffer = []
            self._buffer_rows = 0

        if to_write.height == 0:
            return

        out_path = get_unique_save_path(f"{self.base_path_no_ext}_part{self.part_number}_OFAC_OUTPUT.csv")
        to_write.write_csv(out_path)
        self.written_parts.append((out_path, to_write.height))
        self.part_number += 1

    def finalize(self):
        self._flush(up_to_rows=None)
        return self.written_parts


# ==================== TEXT FILE PROCESSING ====================

def _read_text_chunks(file_path, encoding, delimiter, quotechar, chunk_rows=CHUNK_ROWS):
    """
    THE FIX for bug #1: correctly re-calls next_batches() every loop
    iteration (a BatchedCsvReader has no __iter__, so `for batch in reader`
    -- what v2 tried -- raises TypeError; and forgetting to reassign
    `batches` inside the loop -- what v1 did -- is an infinite loop).
    """
    reader = pl.read_csv_batched(
        file_path, has_header=False, encoding=encoding,
        separator=delimiter, quote_char=quotechar, batch_size=chunk_rows,
    )
    batches = reader.next_batches(1)
    while batches:
        for batch_df in batches:
            yield batch_df
        batches = reader.next_batches(1)


def process_text_file(job, file_path, file_name, header_sets):
    file_size = os.path.getsize(file_path)
    encoding, delimiter, quotechar = detect_encoding_and_dialect(file_path)

    if file_size < TEXT_SIZE_TO_CHUNK:
        df = pl.read_csv(file_path, has_header=False, encoding=encoding, separator=delimiter, quote_char=quotechar)
        _process_dataframe_and_log(job, df, file_path, file_name, "", header_sets, row_count_hint=df.height)
        return

    # Chunked path
    base_path = os.path.join(job.csv_folder, file_name)
    writer = PartedCSVWriter(base_path)
    first_match = None
    total_missing_name = 0
    total_row_count = 0
    detection_meta = None

    for batch_df in _read_text_chunks(file_path, encoding, delimiter, quotechar):
        if job.stop_flag():
            break
        total_row_count += batch_df.height

        if first_match is None:
            sample_rows = batch_df.head(min(batch_df.height, MAX_SEARCH_ROWS)).rows()
            matches = detection.detect_all(sample_rows, batch_df.width, header_sets, max_search_rows=MAX_SEARCH_ROWS)
            if not matches:
                db.log_error(job.run_id, file_path, file_name, ValueError("No identifiable headers or content pattern found"))
                return
            first_match = matches[0]  # text files: one table assumed, no side-by-side multi-table concept
            detection_meta = _match_to_metadata(batch_df, first_match, "")
            output_df, missing = engine.build_output_df(batch_df, first_match, job.company_code, file_path, "")
        else:
            # Later chunks have no header row -- reuse the same column
            # mapping, but tell build_output_df not to skip any rows.
            continuation_match = dict(first_match)
            continuation_match["header_row_idx"] = -1
            output_df, missing = engine.build_output_df(batch_df, continuation_match, job.company_code, file_path, "")

        total_missing_name += missing
        writer.add(output_df)

    written_parts = writer.finalize()
    _write_file_log_for_parts(job, file_path, file_name, "", detection_meta, written_parts, total_row_count, total_missing_name)


def _process_dataframe_and_log(job, df, file_path, file_name, sheet_name, header_sets, row_count_hint,
                                password_matched=None, password_label=None):
    """Shared path for 'small enough to load whole' text and Excel sheets."""
    output_df, metadata_list = engine.extract_from_dataframe(df, header_sets, job.company_code, file_path, sheet_name)

    if not metadata_list:
        db.log_error(job.run_id, file_path, file_name, ValueError("No identifiable headers or content pattern found"), sheet_name=sheet_name)
        return

    for meta in metadata_list:
        table_output = output_df.filter(pl.col("SHEET") == meta["sheet_name"]) if len(metadata_list) > 1 else output_df
        base_path = os.path.join(job.csv_folder, f"{file_name}[{meta['sheet_name']}]" if meta["sheet_name"] else file_name)
        writer = PartedCSVWriter(base_path)
        writer.add(table_output)
        written_parts = writer.finalize()
        _write_file_log_for_parts(
            job, file_path, file_name, meta["sheet_name"], meta, written_parts,
            row_count_hint, meta["missing_name_count"],
            password_matched=password_matched, password_label=password_label,
        )


def _match_to_metadata(df, match, sheet_name):
    name_cols_used = [df.columns[i] for i in match["columns"].get("name", [])]
    firstlast_cols_used = [df.columns[i] for i in match["columns"].get("firstlastname", [])]
    return {
        "sheet_name": sheet_name,
        "header_row_idx": match["header_row_idx"],
        "content_based": match.get("content_based", False),
        "identified_headers": ", ".join(str(v) for v in match.get("raw_header_values", []) if v),
        "multiple_name": len(name_cols_used) >= 2 or len(firstlast_cols_used) > 2,
        "full_name_header": ", ".join(name_cols_used) if name_cols_used else None,
        "first_last_name_header": ", ".join(firstlast_cols_used) if firstlast_cols_used else None,
        "sex_header": df.columns[match["columns"]["sex"][0]] if match["columns"].get("sex") else None,
        "dob_header": df.columns[match["columns"]["dob"][0]] if match["columns"].get("dob") else None,
        "policy_number_header": df.columns[match["columns"]["policynum"][0]] if match["columns"].get("policynum") else None,
    }


def _write_file_log_for_parts(job, file_path, file_name, sheet_name, meta, written_parts, row_count, missing_name_count,
                               password_matched=None, password_label=None):
    if not written_parts:
        db.write_file_log(
            job.run_id, file_path=file_path, file_name=file_name, sheet_name=sheet_name,
            identified_headers=meta.get("identified_headers") if meta else None,
            row_count=row_count, output_row_count=0,
            password_matched=password_matched, password_label=password_label,
            remarks=f"No output rows (missing_name_count={missing_name_count})",
        )
        return

    for idx, (out_path, out_row_count) in enumerate(written_parts, start=1):
        part_suffix = f" (part {idx}/{len(written_parts)})" if len(written_parts) > 1 else ""
        db.write_file_log(
            job.run_id,
            file_path=file_path, file_name=file_name,
            extension=os.path.splitext(file_name)[1].lstrip("."),
            sheet_name=sheet_name,
            password_matched=password_matched, password_label=password_label,
            identified_headers=meta.get("identified_headers") if meta else None,
            multiple_name=meta.get("multiple_name") if meta else None,
            row_count=row_count,
            output_row_count=out_row_count,
            output_csv_path=out_path,
            first_last_name_header=meta.get("first_last_name_header") if meta else None,
            full_name_header=meta.get("full_name_header") if meta else None,
            policy_number_header=meta.get("policy_number_header") if meta else None,
            dob_header=meta.get("dob_header") if meta else None,
            sex_header=meta.get("sex_header") if meta else None,
            remarks=f"missing_name_count={missing_name_count}{part_suffix}",
        )


# ==================== EXCEL FILE PROCESSING ====================

def unlock_excel_file(file_path, password_entries):
    """
    Returns (readable_path_or_bytesio, label_used_or_None). label is None if
    the file wasn't encrypted at all. Raises if encrypted and no password
    in password_entries works.
    """
    with open(file_path, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        if not office_file.is_encrypted():
            return file_path, None

        def attempt(password):
            office_file.load_key(password=password)
            decrypted = io.BytesIO()
            office_file.decrypt(decrypted)
            decrypted.seek(0)
            attempt.result = decrypted
            return True

        result = try_passwords(password_entries, attempt)
        if result is None:
            raise ValueError(f"Could not decrypt {file_path} with any of the {len(password_entries)} provided passwords")
        label, _password = result
        return attempt.result, label


def _read_excel_sheet_chunks(file_path, sheet_name, chunk_rows=CHUNK_ROWS):
    """
    THE FIX for bug #2: Excel files above EXCEL_SIZE_TO_CHUNK now actually
    stream instead of loading the whole workbook into memory. polars has no
    direct equivalent of read_csv_batched for Excel, so this reads rows
    directly via openpyxl's read_only mode (which streams rather than
    materializing the whole sheet) and yields them as small polars
    DataFrames, chunk_rows at a time.
    """
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        buffer = []
        for row in worksheet.iter_rows(values_only=True):
            buffer.append(row)
            if len(buffer) >= chunk_rows:
                yield _rows_to_dataframe(buffer)
                buffer = []
        if buffer:
            yield _rows_to_dataframe(buffer)
    finally:
        workbook.close()


def _rows_to_dataframe(rows):
    width = max(len(r) for r in rows)
    padded = [list(r) + [None] * (width - len(r)) for r in rows]
    columns = [f"column_{i}" for i in range(width)]
    return pl.DataFrame(padded, schema=columns, orient="row")


def process_excel_file(job, file_path, file_name, header_sets, password_entries):
    try:
        readable, matched_password = unlock_excel_file(file_path, password_entries)
    except ValueError as e:
        db.log_error(job.run_id, file_path, file_name, e)
        return
    # Masked immediately, before this touches anything that gets logged --
    # the vault itself is unencrypted now (see ofac_password_vault.py), but
    # the run log is more likely to be shared/exported, so it still never
    # holds the raw value. See mask_password()'s docstring for the format.
    password_label = mask_password(matched_password) if matched_password is not None else None

    file_size = os.path.getsize(file_path) if isinstance(readable, str) else readable.getbuffer().nbytes

    if file_size < EXCEL_SIZE_TO_CHUNK:
        sheets = pl.read_excel(readable, has_header=False, sheet_id=0, raise_if_empty=False)
        for sheet_name, df in sheets.items():
            if job.stop_flag():
                break
            _process_dataframe_and_log(
                job, df, file_path, file_name, sheet_name, header_sets, row_count_hint=df.height,
                password_matched=password_label is not None, password_label=password_label,
            )
        return

    # Chunked path -- need the sheet names first; openpyxl read_only still
    # lets us list sheet names cheaply without loading all rows.
    workbook = openpyxl.load_workbook(readable, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    workbook.close()

    for sheet_name in sheet_names:
        if job.stop_flag():
            break
        base_path = os.path.join(job.csv_folder, f"{file_name}[{sheet_name}]")
        writer = PartedCSVWriter(base_path)
        first_match = None
        detection_meta = None
        total_missing_name = 0
        total_row_count = 0

        for batch_df in _read_excel_sheet_chunks(readable, sheet_name):
            if job.stop_flag():
                break
            total_row_count += batch_df.height

            if first_match is None:
                sample_rows = batch_df.head(min(batch_df.height, MAX_SEARCH_ROWS)).rows()
                matches = detection.detect_all(sample_rows, batch_df.width, header_sets, max_search_rows=MAX_SEARCH_ROWS)
                if not matches:
                    db.log_error(job.run_id, file_path, file_name, ValueError("No identifiable headers or content pattern found"), sheet_name=sheet_name)
                    break
                first_match = matches[0]
                detection_meta = _match_to_metadata(batch_df, first_match, sheet_name)
                output_df, missing = engine.build_output_df(batch_df, first_match, job.company_code, file_path, sheet_name)
            else:
                continuation_match = dict(first_match)
                continuation_match["header_row_idx"] = -1
                output_df, missing = engine.build_output_df(batch_df, continuation_match, job.company_code, file_path, sheet_name)

            total_missing_name += missing
            writer.add(output_df)

        written_parts = writer.finalize()
        if detection_meta is not None:
            _write_file_log_for_parts(
                job, file_path, file_name, sheet_name, detection_meta, written_parts, total_row_count, total_missing_name,
                password_matched=password_label is not None, password_label=password_label,
            )


# ==================== ARCHIVE FILE PROCESSING ====================

def process_archive_file(job, archive_path, archive_name, password_entries):
    files_before = get_all_files(job.unzipped_folder)

    def attempt(password):
        args = [SEVEN_ZIP_PATH, "x", archive_path, "-aou", f"-o{job.unzipped_folder}"]
        if password:
            args.append(f"-p{password}")
        result = subprocess.run(args, capture_output=True, text=True)
        return result.returncode == 0

    # Some archives aren't password protected at all -- try that first.
    result = try_passwords([("(no password)", None)] + password_entries, attempt)

    if result is None:
        db.log_error(job.run_id, archive_path, archive_name, ValueError(f"Extraction failed with all {len(password_entries)} provided passwords"))
        return []

    label, matched_password = result
    files_after = get_all_files(job.unzipped_folder)
    extracted_files = sorted(files_after - files_before)

    password_matched = label != "(no password)"
    # Same masking as the Excel path -- see process_excel_file for why.
    logged_password_label = mask_password(matched_password) if password_matched else None

    db.write_file_log(
        job.run_id, file_path=archive_path, file_name=archive_name,
        extension=os.path.splitext(archive_name)[1].lstrip("."),
        password_matched=password_matched, password_label=logged_password_label,
        remarks=f"Extracted {len(extracted_files)} file(s)",
    )
    return extracted_files


# ==================== TOP-LEVEL DISPATCH ====================

def process_one_file(job, file_path, file_name, header_sets, password_entries):
    ext = os.path.splitext(file_name)[1].lower().lstrip(".")

    if ext in FILE_EXTENSIONS_EXCEL:
        process_excel_file(job, file_path, file_name, header_sets, password_entries)
        return []
    elif ext in FILE_EXTENSIONS_TEXT:
        process_text_file(job, file_path, file_name, header_sets)
        return []
    elif ext in FILE_EXTENSIONS_ARCHIVE:
        return process_archive_file(job, file_path, file_name, password_entries)
    else:
        return []  # unsupported extension -- caller should have filtered already


def run_scan_job(job, file_paths, header_sets, password_entries, progress_callback=None):
    """
    file_paths: initial list of full paths to process (archive contents get
    appended to this as they're extracted, same pattern as v1/v2).
    Files are archived only after they finish processing without raising --
    a failed file stays in place rather than disappearing into Archived/.
    """
    files_processed = 0
    files_failed = 0
    idx = 0

    while idx < len(file_paths):
        if job.stop_flag():
            break

        file_path = file_paths[idx]
        file_name = os.path.basename(file_path)
        if progress_callback:
            progress_callback(idx + 1, len(file_paths), file_name)

        try:
            extra_files = process_one_file(job, file_path, file_name, header_sets, password_entries)
            file_paths.extend(extra_files)
            move_file_to_archived(file_path, job.archived_folder)
            files_processed += 1
        except Exception as e:
            db.log_error(job.run_id, file_path, file_name, e)
            files_failed += 1

        idx += 1

    return files_processed, files_failed
